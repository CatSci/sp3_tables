import pandas as pd
import streamlit as st
import numpy as np
from collections import OrderedDict
import math
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
import openpyxl


st.markdown('**Note - Please do not post target or intermediate structure information externally**.')
st.title('SP3 Table')

uploaded_file = st.file_uploader("Choose a file")


def read_rt(file, sheet_name):
    """Return a list

    extract all retention time from different time reaction
    """
    rt_list = []
    for i in sheet_name:
        df = pd.read_excel(file , sheet_name = i)
        rt_col = df['RT']
        rt_list.append(rt_col)
    
    rt = [i for sub_list in rt_list for i in sub_list]
    return rt

def add_rrt(dataframe):
    """Return a dataframe

    add rrt column in dataframe
    """
    max_area = dataframe['Total Area %'].max()
    rt_max = dataframe['Total Area %'].idxmax()
    dataframe['RRT'] = dataframe['RT'] / dataframe['RT'].iloc[rt_max]
    
    return dataframe

def fill_rrt(file, sheets, df):
    """Returns a dataframe

    fill the rrt value in dataframe
    """
    for i in sheets:
        d = pd.read_excel(file, sheet_name = i)
        row = d.shape[0]
        d = add_rrt(d)        # dataframe with rrt column
        for x in range(d.shape[0]):
            rrt_val = d.loc[x, 'RRT']  # rrt_value
            rt_val = d.loc[x, 'RT']
            idx = df.index[(df["RT"]== rt_val)]     # find index position for rt value in sp3
            df.iloc[idx, 1] = rrt_val

    return df

def add_data(file, sheets, dataframe):
    """Return a dataframe

    add data from different screen reaction time
    """
    
    for i in sheets:
        d = pd.read_excel(file, sheet_name = i)
        dataframe[i] = pd.DataFrame([[np.nan]], index = dataframe.index)
        # st.write(i)
        # st.write(dataframe)
        for idx, val in d.iterrows():
            area, rt_val = val['Total Area %'], val['RT']
            dataframe.at[dataframe.index[dataframe.RT == rt_val].tolist()[0], str(i)] = area
            # st.write(dataframe)
    
    return dataframe


def find_same_col(df_file):
    """Return a dictionary

    to find if the two nearby values of retention time is almost same or not
    if its the same we need to drop it
    """
    SP3 = df_file
    yes = {}
    for i in range(SP3.shape[1] - 2):
        x = SP3.iloc[0, i + 1]
        y = SP3.iloc[0, i + 2]
        if y - x <= 0.05:
            yes[i + 2] = y
        else:
            pass

    return yes

def drop_col(df, rt_val = 16.69, col = 113): 
    """Return a dataframe and the values of the column we need to drop

    to extract similar column values and drop the column
    """
    a = {}
    for i in range(df.shape[0] - 1):
        p = df.iloc[i + 1, col]
        t = i + 1
        a[t] = p
    
    # drop that column
    df.drop([col - 1], axis = 1, inplace = True)
    return df , a


def replace_val(df, drop_col_val, col = 112):
    """Return a dataframe

    to replace the value in nearby column 
    """
    for i in range(df.shape[0] - 1):
        if math.isnan(drop_col_val[i + 1]):
            pass
        else:
            df.iloc[i + 1, col] = drop_col_val[i + 1]
    return df

def sp3_table(df, rev):
    """Returns a dataframe

    columns will be drop and then values will be replaced
    """
    for key, value in rev.items():
        st.write('before dropping')
        st.write(df)
        new_df, a = drop_col(df = df, rt_val = value, col = key)
        st.write('a values')
        st.write(a)
        st.write('after dropping')
        st.write(new_df)
        final_df = replace_val(df = new_df, drop_col_val = a, col = key - 1)
    
    return final_df

def variables():
    col1, col2 = st.columns(2)
    with col1:
        start_retention = st.number_input("Retention time starting value", step= 0.02, min_value= 0.0)
    with col2:
        last_retention = st.number_input('Retention time last value', step= 0.02, min_value= 0.0)
    
    area_threshold = st.number_input("Total area % threshold", step = 0.5, min_value= 0.0)
    
    return start_retention, last_retention, area_threshold

def filter_data(dataframe: pd.DataFrame, start_retention: float, last_retention: float, area_threshold: float):
    """_summary_

    Args:
        dataframe (pd.DataFrame): _description_
        start_retention (float): _description_
        last_retention (float): _description_

    Returns:
        _type_: _description_
    """
    filtered_df = dataframe[(dataframe['RT'] >= start_retention) & (dataframe['RT'] <= last_retention)]
    
    columns = list(filtered_df.columns)
    over_threshold_df = pd.DataFrame(columns=columns)  # DataFrame for rows that satisfy the condition
    under_threshold_df = pd.DataFrame(columns=columns)  # DataFrame for rows that don't satisfy the condition
    
    for r in range(filtered_df.shape[0]):
        row = filtered_df.iloc[r]
        # st.write(row)
        if any(row[c] > area_threshold for c in columns if c not in ['RT', 'RRT']):
            over_threshold_df = over_threshold_df.append(row)
        else:
            under_threshold_df = under_threshold_df.append(row)
    # Drop rows in u_df that are also in f_df
    under_threshold_df = under_threshold_df[~under_threshold_df.apply(tuple, axis=1).isin(over_threshold_df.apply(tuple, axis=1))].reset_index(drop=True)
    
    # Reset the index for both dataframes
    over_threshold_df.reset_index(drop=True, inplace=True)
    
    
    return over_threshold_df, under_threshold_df

if uploaded_file is not None:
    start_retention, last_retention, area_threshold = variables()
def create_sp3_table(): 
    
    # if uploaded_file is not None and start_retention is not None and last_retention is not None and area_threshold is not None:
    if uploaded_file is not None:
        wb = openpyxl.load_workbook(uploaded_file)
        sheets = wb.sheetnames


        wb = openpyxl.load_workbook(uploaded_file)
        sheets = wb.sheetnames
        
        rt_list = read_rt(uploaded_file, sheets)
        rt_list.sort()
        SP3 = pd.DataFrame({'RT' : rt_list, 'RRT': np.nan})


        SP3 = fill_rrt(uploaded_file, sheets, df = SP3)
        SP3 = add_data(uploaded_file, sheets, dataframe = SP3)

        over_threshold_df, under_threshold_df = filter_data(dataframe= SP3, start_retention= start_retention, last_retention= last_retention, area_threshold= area_threshold)

        over_threshold_df = over_threshold_df.T
        over_threshold_df.reset_index(inplace=True)
        under_threshold_df = under_threshold_df.T
        under_threshold_df.reset_index(inplace= True)

        # Create an Excel writer object
        excel_writer = pd.ExcelWriter("output.xlsx", engine="xlsxwriter")

        # Write DataFrames to different sheets
        over_threshold_df.to_excel(excel_writer, sheet_name="Over_Threshold", index=False)
        under_threshold_df.to_excel(excel_writer, sheet_name="Under_Threshold", index=False)

        # Save and close the Excel writer
        excel_writer.save()

        # Create a download link for the generated Excel file
        st.markdown(f"**Download Excel File:**")
        with open("output.xlsx", "rb") as file:
            st.download_button("Download Excel", file.read(), "output.xlsx", key="download")

m = st.markdown("""
<style>
div.stButton > button:first-child {
    background-color: #0C1B2A;
    color:#ffffff;
    border:None;
}

</style>""", unsafe_allow_html=True)

if st.button('Create SP3 Table'):
    create_sp3_table()