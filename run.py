import pandas as pd
import streamlit as st
import numpy as np
from collections import OrderedDict
import math
import openpyxl

from main_utils import read_rt, fill_rrt, add_data, filter_data, compress_data


# hide streamlit style
hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)


st.markdown('<link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" integrity="sha384-Gn5384xqQ1aoWXA+058RXPxPg6fy4IWvTNh0E263XmFcJlSAwiGgFAW/dAiS6JXm" crossorigin="anonymous">', unsafe_allow_html=True)
# background-color: #ed9439;
st.markdown("""
<style>
.navbar {
  height: 80px;
  background-color: #ed9439;
  color: #ed9439;
}
.navbar-brand{
    font-size: 40px;
    margin-left:40px;
}
</style>""", unsafe_allow_html= True)


st.markdown("""
<nav class="navbar fixed-top navbar-expand-lg navbar-dark">
  <a class="navbar-brand" href="#" target="_blank">CatSci</a>
  

</nav>
""", unsafe_allow_html=True)



st.markdown("""
<style>
div.stButton > button:first-child {
    background-color: #ed9439;
    color:#0f1b2a;
    border:None;
}
div.stButton > button:first-child:focus {
    background-color: #ed9439;
    color:#0f1b2a;
    border:None;
}
</style>""", unsafe_allow_html=True)


st.title('SP3 Table')
st.info('Note - Please do not post target or intermediate structure information externally.')

uploaded_file = st.file_uploader("Choose a file")


def variables():
    col1, col2 = st.columns(2)
    with col1:
        start_retention = st.number_input("Retention time starting value", step= 0.02, min_value= 0.0)
    with col2:
        last_retention = st.number_input('Retention time last value', value = 100.00, step= 0.02, min_value= 0.0)
    
    area_threshold = st.number_input("Total area % threshold", step = 0.5, min_value= 0.0)
    
    return start_retention, last_retention, area_threshold




if uploaded_file is not None:
    start_retention, last_retention, area_threshold = variables()
    data_minimize = st.checkbox('Compress Data')
    if data_minimize:
        st.warning('Some Data will be lost', icon="⚠️")



def create_sp3_table(): 
    
    # if uploaded_file is not None and start_retention is not None and last_retention is not None and area_threshold is not None:
    if uploaded_file is not None:
        wb = openpyxl.load_workbook(uploaded_file)
        sheets = wb.sheetnames

        wb = openpyxl.load_workbook(uploaded_file)
        sheets = wb.sheetnames
        
        rt_list = read_rt(uploaded_file, sheets)
        rt_list.sort()
        sp3_df = pd.DataFrame({'RT' : rt_list, 'RRT': np.nan})

        sp3_df = fill_rrt(uploaded_file, sheets, df = sp3_df)
        sp3_df = add_data(uploaded_file, sheets, dataframe = sp3_df)

        over_threshold_df, under_threshold_df = filter_data(dataframe= sp3_df, start_retention= start_retention, last_retention= last_retention, area_threshold= area_threshold)

        over_threshold_df = over_threshold_df.T
        over_threshold_df.reset_index(inplace=True)
        under_threshold_df = under_threshold_df.T
        under_threshold_df.reset_index(inplace= True)

        if data_minimize:
            over_df = compress_data(dataframe=over_threshold_df) if over_threshold_df.shape[1] > 1 else over_threshold_df
            under_df = compress_data(dataframe=under_threshold_df) if under_threshold_df.shape[1] > 1 else under_threshold_df
        
        else:
            over_df = over_threshold_df
            under_df = under_threshold_df
        
        
        # Create an Excel writer object
        excel_writer = pd.ExcelWriter("output.xlsx", engine="openpyxl")

        # Write DataFrames to different sheets
        over_df.to_excel(excel_writer, sheet_name="Over_Threshold", index=False)
        under_df.to_excel(excel_writer, sheet_name="Under_Threshold", index=False)

        # Save and close the Excel writer
        # excel_writer.save()
        excel_writer.close()

        # Create a download link for the generated Excel file
        st.markdown(f"**Download Excel File:**")
        with open("output.xlsx", "rb") as file:
            st.download_button("Download Excel", file.read(), "output.xlsx", key="download")


if st.button('Create SP3 Table'):
    with st.spinner('Working...'):
        create_sp3_table()



